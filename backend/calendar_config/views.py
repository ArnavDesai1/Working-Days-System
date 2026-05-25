from openpyxl import load_workbook
from rest_framework import parsers, permissions, status, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response

from audit_logs.utils import record_audit
from backend.conflicts import conflict_response, is_stale
from .models import Holiday, WorkingDayConfig
from .serializers import HolidaySerializer, WorkingDayConfigSerializer
from users.permissions import CalendarSetupWritePermission


MONTH_LOOKUP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

WEEKDAY_ALIASES = {
    "mon": "mon",
    "monday": "mon",
    "tue": "tue",
    "tues": "tue",
    "tuesday": "tue",
    "wed": "wed",
    "wednesday": "wed",
    "thu": "thu",
    "thur": "thu",
    "thurs": "thu",
    "thursday": "thu",
    "fri": "fri",
    "friday": "fri",
    "sat": "sat",
    "saturday": "sat",
    "sun": "sun",
    "sunday": "sun",
}

FIELD_ALIASES = {
    "client": "client_name",
    "clientname": "client_name",
    "company": "client_name",
    "companyname": "client_name",
    "year": "year",
    "month": "month",
    "workingdays": "working_days",
    "workingday": "working_days",
    "billabledays": "working_days",
    "billabledayscount": "working_days",
    "totalworkingdays": "working_days",
    "weekendpolicy": "weekend_policy",
    "weekend": "weekend_policy",
}


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def compact_key(value):
    return "".join(character.lower() for character in clean_cell(value) if character.isalnum())


def to_int(value):
    text = clean_cell(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        digits = "".join(character for character in text if character.isdigit())
        return int(digits) if digits else None


def parse_month(value):
    text = clean_cell(value).lower()
    if not text:
        return None
    number = to_int(text)
    if number and 1 <= number <= 12:
        return number
    for token in text.replace("-", " ").replace("/", " ").split():
        if token in MONTH_LOOKUP:
            return MONTH_LOOKUP[token]
    return MONTH_LOOKUP.get(text[:3])


def parse_bool(value):
    text = clean_cell(value).lower()
    if text in {"yes", "y", "true", "1", "working", "workday", "on", "checked"}:
        return True
    if text in {"no", "n", "false", "0", "off", "holiday", "weekend", "unchecked"}:
        return False
    return None


def parse_weekend_policy(value):
    text = clean_cell(value).lower()
    if "unpaid" in text:
        return "unpaid"
    if "paid" in text:
        return "paid"
    return None


def workbook_from_upload(uploaded_file):
    uploaded_file.seek(0)
    return load_workbook(uploaded_file, read_only=True, data_only=True)


def sheet_preview(sheet, row_limit=20, col_limit=10):
    rows = []
    for row in sheet.iter_rows(max_row=row_limit, max_col=col_limit, values_only=True):
        rows.append([clean_cell(cell) for cell in row])
    while rows and not any(rows[-1]):
        rows.pop()
    return rows


def extract_calendar_fields(workbook, target_sheet=None):
    fields = {}
    sources = {}
    preview_rows = {}
    holidays = []

    sheets_to_process = workbook.worksheets
    if target_sheet:
        sheets_to_process = [s for s in workbook.worksheets if s.title == target_sheet] or workbook.worksheets

    for sheet in sheets_to_process:
        rows = list(sheet.iter_rows(max_row=80, max_col=20, values_only=True))
        preview_rows[sheet.title] = [[clean_cell(cell) for cell in row[:10]] for row in rows[:20]]
        holiday_header_row = None

        for row_index, row in enumerate(rows):
            for col_index, cell in enumerate(row):
                key = compact_key(cell)
                if not key:
                    continue

                # Detect holiday table header
                if key in {"holidays", "holiday", "holidaylist", "holidayschedule"}:
                    holiday_header_row = row_index
                    continue

                field = FIELD_ALIASES.get(key)
                weekday = WEEKDAY_ALIASES.get(key)
                right_value = row[col_index + 1] if col_index + 1 < len(row) else None
                below_value = rows[row_index + 1][col_index] if row_index + 1 < len(rows) and col_index < len(rows[row_index + 1]) else None
                candidate_values = [right_value, below_value]

                if field and field not in fields:
                    for candidate in candidate_values:
                        if field == "month":
                            parsed = parse_month(candidate)
                        elif field in {"year", "working_days"}:
                            parsed = to_int(candidate)
                        elif field == "weekend_policy":
                            parsed = parse_weekend_policy(candidate)
                        else:
                            parsed = clean_cell(candidate)
                        if parsed not in {"", None}:
                            fields[field] = parsed
                            sources[field] = f"{sheet.title}!R{row_index + 1}C{col_index + 1}"
                            break

                if weekday and weekday not in fields:
                    parsed = None
                    for candidate in candidate_values:
                        parsed = parse_bool(candidate)
                        if parsed is not None:
                            break
                    if parsed is not None:
                        fields[weekday] = parsed
                        sources[weekday] = f"{sheet.title}!R{row_index + 1}C{col_index + 1}"

                if "month" not in fields:
                    parsed_month = parse_month(cell)
                    if parsed_month:
                        fields["month"] = parsed_month
                        sources["month"] = f"{sheet.title}!R{row_index + 1}C{col_index + 1}"
                if "year" not in fields:
                    parsed_year = to_int(cell)
                    if parsed_year and 2000 <= parsed_year <= 2100:
                        fields["year"] = parsed_year
                        sources["year"] = f"{sheet.title}!R{row_index + 1}C{col_index + 1}"

        # Extract holidays from the table below the "Holidays" marker
        if holiday_header_row is not None:
            _extract_holiday_rows(rows, holiday_header_row, holidays)

    relevant_fields = ["client_name", "year", "month", "working_days", "mon", "tue", "wed", "thu", "fri", "sat", "sun", "weekend_policy"]
    matched = [field for field in relevant_fields if field in fields]
    missing = [field for field in relevant_fields if field not in fields]
    return {
        "valid": bool(matched),
        "fields": fields,
        "matched_fields": matched,
        "missing_fields": missing,
        "sources": sources,
        "preview_rows": preview_rows,
        "holidays": holidays,
        "message": "Relevant calendar fields were extracted." if matched else "No relevant calendar fields found in this Excel file.",
    }


HOLIDAY_COLUMN_ALIASES = {
    "holidayname": "name",
    "holiday": "name",
    "name": "name",
    "date": "date",
    "holidaydate": "date",
    "type": "type",
    "holidaytype": "type",
    "category": "type",
    "duration": "duration_days",
    "durationdays": "duration_days",
    "days": "duration_days",
    "noofdays": "duration_days",
}


def _parse_date_cell(value):
    """Parse a date from an Excel cell — handles datetime objects and string formats."""
    from datetime import date, datetime

    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_cell(value)
    if not text:
        return ""
    # Try common date formats
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _extract_holiday_rows(rows, marker_row_index, holidays):
    """Extract holiday entries from rows below the holiday marker."""
    # The row after the marker should be column headers (Holiday Name | Date | Type | Duration)
    # Or the marker row itself might contain the column headers in adjacent cells
    header_row_index = marker_row_index + 1
    if header_row_index >= len(rows):
        return

    header_row = rows[header_row_index]
    column_mapping = {}
    for col_index, cell in enumerate(header_row):
        key = compact_key(cell)
        if key in HOLIDAY_COLUMN_ALIASES:
            column_mapping[HOLIDAY_COLUMN_ALIASES[key]] = col_index

    # If we didn't find at least a "name" column, try the marker row itself as headers
    if "name" not in column_mapping:
        header_row = rows[marker_row_index]
        column_mapping = {}
        for col_index, cell in enumerate(header_row):
            key = compact_key(cell)
            if key in HOLIDAY_COLUMN_ALIASES:
                column_mapping[HOLIDAY_COLUMN_ALIASES[key]] = col_index
        header_row_index = marker_row_index

    if "name" not in column_mapping:
        return

    # Parse data rows after the header
    for data_row_index in range(header_row_index + 1, len(rows)):
        data_row = rows[data_row_index]
        if not any(data_row):
            break  # Stop at first fully blank row

        name = clean_cell(data_row[column_mapping["name"]]) if "name" in column_mapping and column_mapping["name"] < len(data_row) else ""
        date_val = _parse_date_cell(data_row[column_mapping["date"]]) if "date" in column_mapping and column_mapping["date"] < len(data_row) else ""
        holiday_type = clean_cell(data_row[column_mapping["type"]]).lower() if "type" in column_mapping and column_mapping["type"] < len(data_row) else "public"
        duration = to_int(data_row[column_mapping["duration_days"]]) if "duration_days" in column_mapping and column_mapping["duration_days"] < len(data_row) else 1

        if not name and not date_val:
            continue  # Skip rows where both name and date are empty

        if holiday_type not in {"public", "company"}:
            holiday_type = "public"

        holiday = {
            "name": name,
            "date": date_val,
            "type": holiday_type,
            "duration_days": max(1, duration or 1),
        }

        # Add warnings for incomplete entries
        warnings = []
        if not name:
            warnings.append("Holiday name is missing.")
        if not date_val:
            warnings.append("Holiday date is missing.")
        if warnings:
            holiday["warning"] = " ".join(warnings)

        holidays.append(holiday)



class WorkingDayConfigViewSet(viewsets.ModelViewSet):
    queryset = WorkingDayConfig.objects.select_related("client").all()
    serializer_class = WorkingDayConfigSerializer
    permission_classes = [permissions.IsAuthenticated, CalendarSetupWritePermission]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        existing = WorkingDayConfig.objects.filter(
            client=data["client"],
            year=data["year"],
            month=data["month"],
        ).first()
        if existing and is_stale(existing, request):
            return conflict_response()
        previous_values = None
        if existing:
            previous_values = {
                "working_days": existing.working_days,
                "mon": existing.mon,
                "tue": existing.tue,
                "wed": existing.wed,
                "thu": existing.thu,
                "fri": existing.fri,
                "sat": existing.sat,
                "sun": existing.sun,
                "weekend_policy": existing.weekend_policy,
            }
        instance, _ = WorkingDayConfig.objects.update_or_create(
            client=data["client"],
            year=data["year"],
            month=data["month"],
            defaults={
                "working_days": data["working_days"],
                "mon": data["mon"],
                "tue": data["tue"],
                "wed": data["wed"],
                "thu": data["thu"],
                "fri": data["fri"],
                "sat": data["sat"],
                "sun": data["sun"],
                "weekend_policy": data["weekend_policy"],
            },
        )
        record_audit(
            request.user,
            "CALENDAR_RULES_UPDATED" if existing else "CALENDAR_RULES_CREATED",
            "working_day_config",
            instance.id,
            f"{instance.client.name} - {instance.month}/{instance.year}",
            {
                "client": instance.client.name,
                "year": instance.year,
                "month": instance.month,
                "working_days": instance.working_days,
                "weekend_policy": instance.weekend_policy,
                "weekdays": {
                    "mon": instance.mon,
                    "tue": instance.tue,
                    "wed": instance.wed,
                    "thu": instance.thu,
                    "fri": instance.fri,
                    "sat": instance.sat,
                    "sun": instance.sun,
                },
                "previous_values": previous_values,
            },
        )
        return Response(self.get_serializer(instance).data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_stale(instance, request):
            return conflict_response()
        return super().update(request, *args, **kwargs)

    def perform_update(self, serializer):
        previous_values = {
            "working_days": serializer.instance.working_days,
            "mon": serializer.instance.mon,
            "tue": serializer.instance.tue,
            "wed": serializer.instance.wed,
            "thu": serializer.instance.thu,
            "fri": serializer.instance.fri,
            "sat": serializer.instance.sat,
            "sun": serializer.instance.sun,
            "weekend_policy": serializer.instance.weekend_policy,
        }
        instance = serializer.save()
        record_audit(
            self.request.user,
            "CALENDAR_RULES_UPDATED",
            "working_day_config",
            instance.id,
            f"{instance.client.name} - {instance.month}/{instance.year}",
            {
                "client": instance.client.name,
                "year": instance.year,
                "month": instance.month,
                "working_days": instance.working_days,
                "weekend_policy": instance.weekend_policy,
                "weekdays": {
                    "mon": instance.mon,
                    "tue": instance.tue,
                    "wed": instance.wed,
                    "thu": instance.thu,
                    "fri": instance.fri,
                    "sat": instance.sat,
                    "sun": instance.sun,
                },
                "previous_values": previous_values,
            },
        )


class HolidayViewSet(viewsets.ModelViewSet):
    queryset = Holiday.objects.select_related("client", "created_by").all()
    serializer_class = HolidaySerializer
    permission_classes = [permissions.IsAuthenticated, CalendarSetupWritePermission]

    def perform_create(self, serializer):
        holiday = serializer.save(created_by=self.request.user)
        record_audit(
            self.request.user,
            "HOLIDAY_CREATED",
            "holiday",
            holiday.id,
            holiday.name,
            {
                "client": holiday.client.name,
                "date": str(holiday.date),
                "duration_days": holiday.duration_days,
                "type": holiday.type,
            },
        )

    def perform_update(self, serializer):
        previous = {
            "name": serializer.instance.name,
            "date": str(serializer.instance.date),
            "duration_days": serializer.instance.duration_days,
            "type": serializer.instance.type,
        }
        holiday = serializer.save()
        record_audit(
            self.request.user,
            "HOLIDAY_UPDATED",
            "holiday",
            holiday.id,
            holiday.name,
            {
                "client": holiday.client.name,
                "date": str(holiday.date),
                "duration_days": holiday.duration_days,
                "type": holiday.type,
                "previous_values": previous,
            },
        )

    def perform_destroy(self, instance):
        record_audit(
            self.request.user,
            "HOLIDAY_DELETED",
            "holiday",
            instance.id,
            instance.name,
            {
                "client": instance.client.name,
                "date": str(instance.date),
                "duration_days": instance.duration_days,
                "type": instance.type,
            },
        )
        instance.delete()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_stale(instance, request):
            return conflict_response()
        return super().update(request, *args, **kwargs)


class CalendarExcelPreviewView(APIView):
    parser_classes = [parsers.MultiPartParser]
    permission_classes = [permissions.IsAuthenticated, CalendarSetupWritePermission]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Upload an Excel file."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            workbook = workbook_from_upload(uploaded_file)
        except Exception:
            return Response({"detail": "Invalid Excel file. Upload a readable .xlsx file."}, status=status.HTTP_400_BAD_REQUEST)

        sheets = [
            {
                "name": sheet.title,
                "rows": sheet_preview(sheet),
            }
            for sheet in workbook.worksheets[:5]
        ]
        return Response({"file_name": uploaded_file.name, "sheets": sheets})


class CalendarExcelExtractView(APIView):
    parser_classes = [parsers.MultiPartParser]
    permission_classes = [permissions.IsAuthenticated, CalendarSetupWritePermission]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Upload an Excel file."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            workbook = workbook_from_upload(uploaded_file)
            target_sheet = request.data.get("sheet_name", "")
            extraction = extract_calendar_fields(workbook, target_sheet=target_sheet or None)
        except Exception:
            return Response({"detail": "Invalid Excel file. Upload a readable .xlsx file."}, status=status.HTTP_400_BAD_REQUEST)

        # Include sheet names for frontend sheet selector
        extraction["sheet_names"] = [s.title for s in workbook.worksheets]

        # Client detection — check if extracted client is registered
        from clients.models import Client

        client_name = clean_cell(extraction["fields"].get("client_name", ""))
        if client_name:
            exact = Client.objects.filter(name__iexact=client_name, status="active").first()
            partial = (
                Client.objects.filter(name__icontains=client_name, status="active").first()
                if not exact
                else None
            )
            matched_client = exact or partial
            extraction["client_match"] = {
                "excel_client_name": client_name,
                "matched": bool(matched_client),
                "matched_client_name": matched_client.name if matched_client else None,
                "matched_client_id": matched_client.id if matched_client else None,
                "is_exact": bool(exact),
            }
        else:
            extraction["client_match"] = {
                "excel_client_name": "",
                "matched": False,
                "matched_client_name": None,
                "matched_client_id": None,
                "is_exact": False,
            }

        response_status = status.HTTP_200_OK if extraction["valid"] else status.HTTP_422_UNPROCESSABLE_ENTITY
        extraction["file_name"] = uploaded_file.name
        return Response(extraction, status=response_status)

