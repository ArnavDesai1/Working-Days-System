from calendar import monthrange
from datetime import date, datetime
import email
from email.header import decode_header, make_header
from html.parser import HTMLParser
import imaplib
import os

from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework import parsers, permissions, status, viewsets
from rest_framework.response import Response
from rest_framework.decorators import action
from audit_logs.utils import record_audit
from backend.conflicts import conflict_response, is_stale
from locking import acquire_lock, release_lock, get_lock_status, validate_lock_owner
from users.permissions import EmployeeDeploymentWritePermission

from .models import ClientResource, FileUploadLog, MonthlyWorkEntry, PmoTimesheetData, Timesheet
from .serializers import (
    ClientResourceSerializer,
    MonthlyWorkEntrySerializer,
    PmoTimesheetDataSerializer,
    TimesheetSerializer,
)


PUBLIC_HOLIDAY_RECOMMENDATIONS = {
    "fixed": [
        ("Republic Day", 1, 26),
        ("Maharashtra Day", 5, 1),
        ("Independence Day", 8, 15),
        ("Gandhi Jayanti", 10, 2),
        ("Christmas", 12, 25),
    ],
    2026: [
        ("Holi", 3, 4),
        ("Good Friday", 4, 3),
        ("Eid al-Fitr", 3, 20),
        ("Diwali", 11, 8),
    ],
    2027: [
        ("Holi", 3, 22),
        ("Good Friday", 3, 26),
        ("Eid al-Fitr", 3, 10),
        ("Diwali", 10, 29),
    ],
}


def recommended_holidays_for_month(year, month):
    holidays = []
    for name, fixed_month, day in PUBLIC_HOLIDAY_RECOMMENDATIONS["fixed"]:
        if fixed_month == month:
            holiday_date = date(year, fixed_month, day)
            holidays.append(
                {
                    "name": name,
                    "date": holiday_date.isoformat(),
                    "weekday": holiday_date.strftime("%A"),
                    "type": "public",
                    "source": "recommended",
                }
            )

    for name, variable_month, day in PUBLIC_HOLIDAY_RECOMMENDATIONS.get(year, []):
        if variable_month == month:
            holiday_date = date(year, variable_month, day)
            holidays.append(
                {
                    "name": name,
                    "date": holiday_date.isoformat(),
                    "weekday": holiday_date.strftime("%A"),
                    "type": "public",
                    "source": "recommended",
                }
            )
    return holidays


def calculate_working_days(year, month, weekday_flags, holiday_dates=None):
    holiday_dates = set(holiday_dates or [])
    weekday_keys = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
    total = 0
    days = []
    for day in range(1, monthrange(year, month)[1] + 1):
        current = date(year, month, day)
        weekday_key = weekday_keys[current.weekday()]
        is_selected = bool(weekday_flags.get(weekday_key))
        is_holiday = current.isoformat() in holiday_dates
        if is_selected and not is_holiday:
            total += 1
        days.append(
            {
                "date": current.isoformat(),
                "weekday": current.strftime("%A"),
                "is_selected_workday": is_selected,
                "is_recommended_holiday": is_holiday,
            }
        )
    return total, days


FIELD_ALIASES = {
    "client_name": ["client", "client name", "company", "company name"],
    "employee_code": ["employee code", "emp code", "emp id", "employee id", "resource id"],
    "employee_name": ["employee name", "emp name", "resource name", "consultant name", "name"],
    "primary_skill": ["primary skill", "skill", "technology", "role"],
    "po_number": ["po number", "po no", "po", "purchase order"],
    "start_date": ["start date", "from date", "period start"],
    "end_date": ["end date", "to date", "period end"],
    "timesheet_month": ["timesheet month", "month", "billing month", "month year", "period"],
    "timesheet_year": ["timesheet year", "year"],
    "billing_rate": ["billing rate", "monthly rate", "rate", "billing amount"],
    "leaves_taken": ["leaves taken", "leave taken", "leaves", "lop days", "leave days"],
    "dates_of_leaves": ["dates of leaves", "leave dates", "dates of leave"],
    "compoff_days": ["compoff days", "comp off days", "comp-off days"],
    "compoff_dates": ["compoff dates", "comp off dates", "comp-off dates"],
    "total_leave": ["total leave", "net leave", "net leave days"],
    "pmo_billed_amount": ["pmo billed amount", "billed amount", "invoice amount", "amount"],
}

REQUIRED_IMPORT_FIELDS = ["client_name", "employee_code", "employee_name", "timesheet_month"]


def load_local_env():
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), ".env")
    if not os.path.exists(env_path):
        return

    with open(env_path, "r", encoding="utf-8") as env_file:
        for line in env_file:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_local_env()


class HtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self._current_table = None
        self._current_row = None
        self._current_cell = None
        self._capture_cell = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._current_table = []
        elif tag == "tr" and self._current_table is not None:
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []
            self._capture_cell = True

    def handle_data(self, data):
        if self._capture_cell and self._current_cell is not None:
            text = " ".join(data.split())
            if text:
                self._current_cell.append(text)

    def handle_endtag(self, tag):
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            self._current_row.append(" ".join(self._current_cell).strip())
            self._current_cell = None
            self._capture_cell = False
        elif tag == "tr" and self._current_row is not None and self._current_table is not None:
            if any(self._current_row):
                self._current_table.append(self._current_row)
            self._current_row = None
        elif tag == "table" and self._current_table is not None:
            if self._current_table:
                self.tables.append(self._current_table)
            self._current_table = None


def decoded_header(value):
    if not value:
        return ""
    return str(make_header(decode_header(value)))


def normalize_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def to_int(value, default=0):
    if value in ("", None):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_float(value):
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def build_header_mapping(headers):
    normalized_headers = [normalize_header(header) for header in headers]
    mapping = {}

    for field, aliases in FIELD_ALIASES.items():
        possible_names = {normalize_header(field), *[normalize_header(alias) for alias in aliases]}
        for index, header in enumerate(normalized_headers):
            if header in possible_names:
                mapping[field] = index
                break

    return mapping


def find_header_row(sheet):
    best_row = 1
    best_mapping = {}
    best_headers = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        headers = [normalize_cell(value) for value in row]
        mapping = build_header_mapping(headers)
        if len(mapping) > len(best_mapping):
            best_row = row_number
            best_mapping = mapping
            best_headers = headers

    return best_row, best_headers, best_mapping


def parse_excel(file_obj):
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    sheet = workbook.active
    header_row, headers, mapping = find_header_row(sheet)
    columns = [str(header or "").strip() for header in headers if str(header or "").strip()]
    raw_table_rows = []
    raw_rows = []
    normalized_rows = []

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in ("", None) for value in row):
            continue

        raw_row = {}
        raw_table_row = []
        for index, header in enumerate(headers):
            if str(header or "").strip():
                value = normalize_cell(row[index] if index < len(row) else "")
                raw_row[str(header).strip()] = value
                raw_table_row.append(value)

        normalized = {}
        for field, index in mapping.items():
            normalized[field] = normalize_cell(row[index] if index < len(row) else "")

        for field in FIELD_ALIASES:
            normalized.setdefault(field, "")

        normalized["leaves_taken"] = to_int(normalized["leaves_taken"])
        normalized["compoff_days"] = to_int(normalized["compoff_days"])
        normalized["total_leave"] = to_int(normalized["total_leave"])
        normalized["billing_rate"] = to_float(normalized["billing_rate"])
        normalized["pmo_billed_amount"] = to_float(normalized["pmo_billed_amount"])

        raw_rows.append(raw_row)
        raw_table_rows.append(raw_table_row)
        normalized_rows.append(normalized)

    missing_required = [field for field in REQUIRED_IMPORT_FIELDS if field not in mapping]
    return {
        "sheet_name": sheet.title,
        "columns": columns,
        "field_mapping": {field: columns[index] for field, index in mapping.items() if index < len(columns)},
        "missing_required_fields": missing_required,
        "raw_table": {
            "columns": columns,
            "rows": raw_table_rows[:50],
        },
        "rows": raw_rows[:50],
        "normalized_rows": normalized_rows[:50],
        "total_rows": len(normalized_rows),
    }


def normalize_tabular_rows(headers, data_rows):
    columns = [str(header or "").strip() for header in headers]
    mapping = build_header_mapping(columns)
    normalized_rows = []
    raw_rows = []
    raw_table_rows = []

    for row in data_rows:
        raw_row = {}
        raw_table_row = []
        for index, header in enumerate(columns):
            if header:
                value = normalize_cell(row[index] if index < len(row) else "")
                raw_row[header] = value
                raw_table_row.append(value)

        normalized = {}
        for field, index in mapping.items():
            normalized[field] = normalize_cell(row[index] if index < len(row) else "")

        for field in FIELD_ALIASES:
            normalized.setdefault(field, "")

        normalized["leaves_taken"] = to_int(normalized["leaves_taken"])
        normalized["compoff_days"] = to_int(normalized["compoff_days"])
        normalized["total_leave"] = to_int(normalized["total_leave"])
        normalized["billing_rate"] = to_float(normalized["billing_rate"])
        normalized["pmo_billed_amount"] = to_float(normalized["pmo_billed_amount"])

        raw_rows.append(raw_row)
        raw_table_rows.append(raw_table_row)
        normalized_rows.append(normalized)

    return {
        "columns": [column for column in columns if column],
        "field_mapping": {
            field: columns[index] for field, index in mapping.items() if index < len(columns)
        },
        "missing_required_fields": [field for field in REQUIRED_IMPORT_FIELDS if field not in mapping],
        "raw_table": {
            "columns": [column for column in columns if column],
            "rows": raw_table_rows[:50],
        },
        "rows": raw_rows[:50],
        "normalized_rows": normalized_rows[:50],
        "total_rows": len(normalized_rows),
    }


def extract_html_tables(message):
    html_parts = []

    if message.is_multipart():
        for part in message.walk():
            if part.get_content_type() == "text/html":
                payload = part.get_payload(decode=True)
                if payload:
                    html_parts.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
    elif message.get_content_type() == "text/html":
        payload = message.get_payload(decode=True)
        if payload:
            html_parts.append(payload.decode(message.get_content_charset() or "utf-8", errors="replace"))

    parser = HtmlTableParser()
    for html in html_parts:
        parser.feed(html)

    return parser.tables


def fetch_mail_table_preview(mailbox, subject_filter, month_filter, limit):
    password = os.environ.get("MAIL_APP_PASSWORD", "").replace(" ", "")
    host = os.environ.get("MAIL_HOST", "imap.gmail.com")
    port = int(os.environ.get("MAIL_PORT", "993"))
    username = os.environ.get("MAIL_USERNAME", mailbox)

    if not password:
        raise RuntimeError("MAIL_APP_PASSWORD is not configured in the backend environment.")

    previews = []
    with imaplib.IMAP4_SSL(host, port) as connection:
        connection.login(username, password)
        connection.select("INBOX")

        criteria = ['SUBJECT', f'"{subject_filter or "timesheet"}"']
        status_code, data = connection.search(None, *criteria)
        if status_code != "OK":
            raise RuntimeError("Unable to search mailbox.")

        message_ids = data[0].split()[-limit:]
        for message_id in reversed(message_ids):
            status_code, message_data = connection.fetch(message_id, "(RFC822)")
            if status_code != "OK":
                continue

            raw_message = message_data[0][1]
            message = email.message_from_bytes(raw_message)
            subject = decoded_header(message.get("Subject"))
            sender = decoded_header(message.get("From"))

            if month_filter and month_filter.lower() not in subject.lower():
                continue

            tables = extract_html_tables(message)
            for table_index, table in enumerate(tables):
                if len(table) < 2:
                    continue
                preview = normalize_tabular_rows(table[0], table[1:])
                preview.update(
                    {
                        "source": "email_body",
                        "mailbox": username,
                        "subject": subject,
                        "sender": sender,
                        "message_id": message_id.decode(),
                        "table_index": table_index,
                        "file_name": f"email:{message_id.decode()}:table:{table_index}",
                        "sheet_name": "Email body table",
                    }
                )
                previews.append(preview)

    return previews


class TimesheetViewSet(viewsets.ModelViewSet):
    queryset = Timesheet.objects.select_related("client", "submitted_by").all()
    serializer_class = TimesheetSerializer

    def perform_create(self, serializer):
        serializer.save(submitted_by=self.request.user)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def lock(self, request, pk=None):
        obj = self.get_object()
        success, message, lock_data = acquire_lock(obj, request.user)
        if success:
            record_audit(request.user, "LOCK_ACQUIRED", "timesheet", obj.id, f"Timesheet {obj.id}", {"action": "editing_started"})
            return Response(lock_data, status=status.HTTP_200_OK)
        return Response({'error': message, 'lock': lock_data}, status=status.HTTP_423_LOCKED)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def unlock(self, request, pk=None):
        obj = self.get_object()
        success, message = release_lock(obj, request.user)
        if success:
            record_audit(request.user, "LOCK_RELEASED", "timesheet", obj.id, f"Timesheet {obj.id}", {"action": "editing_completed"})
            return Response({'message': message}, status=status.HTTP_200_OK)
        return Response({'error': message}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def lock_status(self, request, pk=None):
        obj = self.get_object()
        return Response(get_lock_status(obj), status=status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        is_valid, error_response = validate_lock_owner(obj, request.user)
        if not is_valid:
            return error_response
        if is_stale(obj, request):
            return conflict_response()
        return super().update(request, *args, **kwargs)


class ClientResourceViewSet(viewsets.ModelViewSet):
    queryset = ClientResource.objects.select_related("client").all()
    serializer_class = ClientResourceSerializer
    permission_classes = [permissions.IsAuthenticated, EmployeeDeploymentWritePermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        client_id = self.request.query_params.get("client")
        if client_id:
            queryset = queryset.filter(client_id=client_id)
        return queryset

    def perform_create(self, serializer):
        resource = serializer.save()
        record_audit(
            self.request.user,
            "EMPLOYEE_CREATED",
            "client_resource",
            resource.id,
            resource.full_name,
            {
                "client": resource.client.name,
                "employee_code": resource.employee_code,
                "email": resource.email,
                "designation": resource.designation,
                "status": resource.status,
            },
        )

    def perform_update(self, serializer):
        previous = {
            "client": serializer.instance.client.name,
            "employee_code": serializer.instance.employee_code,
            "full_name": serializer.instance.full_name,
            "email": serializer.instance.email,
            "designation": serializer.instance.designation,
            "status": serializer.instance.status,
        }
        resource = serializer.save()
        record_audit(
            self.request.user,
            "EMPLOYEE_UPDATED",
            "client_resource",
            resource.id,
            resource.full_name,
            {
                "client": resource.client.name,
                "employee_code": resource.employee_code,
                "email": resource.email,
                "designation": resource.designation,
                "status": resource.status,
                "previous_values": previous,
            },
        )

    def perform_destroy(self, instance):
        record_audit(
            self.request.user,
            "EMPLOYEE_DELETED",
            "client_resource",
            instance.id,
            instance.full_name,
            {
                "client": instance.client.name,
                "employee_code": instance.employee_code,
                "email": instance.email,
            },
        )
        instance.delete()

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def lock(self, request, pk=None):
        obj = self.get_object()
        success, message, lock_data = acquire_lock(obj, request.user)
        if success:
            record_audit(request.user, "LOCK_ACQUIRED", "client_resource", obj.id, obj.full_name, {"action": "editing_started"})
            return Response(lock_data, status=status.HTTP_200_OK)
        return Response({'error': message, 'lock': lock_data}, status=status.HTTP_423_LOCKED)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def unlock(self, request, pk=None):
        obj = self.get_object()
        success, message = release_lock(obj, request.user)
        if success:
            record_audit(request.user, "LOCK_RELEASED", "client_resource", obj.id, obj.full_name, {"action": "editing_completed"})
            return Response({'message': message}, status=status.HTTP_200_OK)
        return Response({'error': message}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def lock_status(self, request, pk=None):
        obj = self.get_object()
        return Response(get_lock_status(obj), status=status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        is_valid, error_response = validate_lock_owner(obj, request.user)
        if not is_valid:
            return error_response
        return super().update(request, *args, **kwargs)


class MonthlyWorkEntryViewSet(viewsets.ModelViewSet):
    queryset = MonthlyWorkEntry.objects.select_related("client", "resource", "submitted_by").all()
    serializer_class = MonthlyWorkEntrySerializer
    permission_classes = [permissions.IsAuthenticated, EmployeeDeploymentWritePermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        client_id = self.request.query_params.get("client")
        year = self.request.query_params.get("year")
        month = self.request.query_params.get("month")
        if client_id:
            queryset = queryset.filter(client_id=client_id)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)
        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        existing = MonthlyWorkEntry.objects.filter(
            resource=data["resource"],
            year=data["year"],
            month=data["month"],
        ).first()
        if existing and is_stale(existing, request):
            return conflict_response()
        instance, _ = MonthlyWorkEntry.objects.update_or_create(
            resource=data["resource"],
            year=data["year"],
            month=data["month"],
            defaults={
                "client": data["client"],
                "expected_working_days": data["expected_working_days"],
                "days_worked": data["days_worked"],
                "leave_days": data["leave_days"],
                "extra_days": data["extra_days"],
                "total_salary": data.get("total_salary"),
                "payable_salary": data.get("payable_salary"),
                "remarks": data.get("remarks", ""),
                "status": data.get("status", "draft"),
                "submitted_by": request.user,
            },
        )
        record_audit(
            request.user,
            "MONTHLY_ENTRY_UPDATED" if existing else "MONTHLY_ENTRY_CREATED",
            "monthly_work_entry",
            instance.id,
            f"{instance.resource.full_name} - {instance.month}/{instance.year}",
            {
                "client": instance.client.name,
                "employee": instance.resource.full_name,
                "employee_code": instance.resource.employee_code,
                "year": instance.year,
                "month": instance.month,
                "expected_working_days": instance.expected_working_days,
                "days_worked": str(instance.days_worked),
                "leave_days": str(instance.leave_days),
                "extra_days": str(instance.extra_days),
                "total_salary": str(instance.total_salary or ""),
                "payable_salary": str(instance.payable_salary or ""),
                "status": instance.status,
            },
        )
        return Response(self.get_serializer(instance).data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        is_valid, error_response = validate_lock_owner(instance, request.user)
        if not is_valid:
            return error_response
        if is_stale(instance, request):
            return conflict_response()
        return super().update(request, *args, **kwargs)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def lock(self, request, pk=None):
        obj = self.get_object()
        success, message, lock_data = acquire_lock(obj, request.user)
        if success:
            record_audit(request.user, "LOCK_ACQUIRED", "monthly_work_entry", obj.id, f"{obj.resource.full_name} - {obj.month}/{obj.year}", {"action": "editing_started"})
            return Response(lock_data, status=status.HTTP_200_OK)
        return Response({'error': message, 'lock': lock_data}, status=status.HTTP_423_LOCKED)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def unlock(self, request, pk=None):
        obj = self.get_object()
        success, message = release_lock(obj, request.user)
        if success:
            record_audit(request.user, "LOCK_RELEASED", "monthly_work_entry", obj.id, f"{obj.resource.full_name} - {obj.month}/{obj.year}", {"action": "editing_completed"})
            return Response({'message': message}, status=status.HTTP_200_OK)
        return Response({'error': message}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def lock_status(self, request, pk=None):
        obj = self.get_object()
        return Response(get_lock_status(obj), status=status.HTTP_200_OK)

    def perform_update(self, serializer):
        previous = {
            "year": serializer.instance.year,
            "month": serializer.instance.month,
            "expected_working_days": serializer.instance.expected_working_days,
            "days_worked": str(serializer.instance.days_worked),
            "leave_days": str(serializer.instance.leave_days),
            "extra_days": str(serializer.instance.extra_days),
            "total_salary": str(serializer.instance.total_salary or ""),
            "payable_salary": str(serializer.instance.payable_salary or ""),
            "status": serializer.instance.status,
            "remarks": serializer.instance.remarks,
        }
        entry = serializer.save()
        record_audit(
            self.request.user,
            "MONTHLY_ENTRY_UPDATED",
            "monthly_work_entry",
            entry.id,
            f"{entry.resource.full_name} - {entry.month}/{entry.year}",
            {
                "client": entry.client.name,
                "employee": entry.resource.full_name,
                "employee_code": entry.resource.employee_code,
                "year": entry.year,
                "month": entry.month,
                "expected_working_days": entry.expected_working_days,
                "days_worked": str(entry.days_worked),
                "leave_days": str(entry.leave_days),
                "extra_days": str(entry.extra_days),
                "total_salary": str(entry.total_salary or ""),
                "payable_salary": str(entry.payable_salary or ""),
                "status": entry.status,
                "remarks": entry.remarks,
                "previous_values": previous,
            },
        )

    def perform_destroy(self, instance):
        record_audit(
            self.request.user,
            "MONTHLY_ENTRY_DELETED",
            "monthly_work_entry",
            instance.id,
            f"{instance.resource.full_name} - {instance.month}/{instance.year}",
            {
                "client": instance.client.name,
                "employee": instance.resource.full_name,
                "employee_code": instance.resource.employee_code,
                "year": instance.year,
                "month": instance.month,
            },
        )
        instance.delete()


class HolidayRecommendationView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = to_int(request.query_params.get("year"), default=datetime.now().year)
        month = to_int(request.query_params.get("month"), default=datetime.now().month)
        weekday_flags = {
            "mon": request.query_params.get("mon", "true") == "true",
            "tue": request.query_params.get("tue", "true") == "true",
            "wed": request.query_params.get("wed", "true") == "true",
            "thu": request.query_params.get("thu", "true") == "true",
            "fri": request.query_params.get("fri", "true") == "true",
            "sat": request.query_params.get("sat", "false") == "true",
            "sun": request.query_params.get("sun", "false") == "true",
        }
        recommended = recommended_holidays_for_month(year, month)
        working_days, calendar_days = calculate_working_days(
            year,
            month,
            weekday_flags,
            [holiday["date"] for holiday in recommended],
        )
        return Response(
            {
                "year": year,
                "month": month,
                "recommended_holidays": recommended,
                "recommended_working_days": working_days,
                "calendar_days": calendar_days,
                "note": "Recommendations are defaults and should be confirmed against the client's official holiday list.",
            }
        )


class PmoTimesheetDataViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PmoTimesheetData.objects.all().order_by("-uploaded_at", "-id")
    serializer_class = PmoTimesheetDataSerializer
    permission_classes = [permissions.IsAuthenticated]






class MailTableFetchView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        mailbox = request.data.get("mailbox", "")
        subject_filter = request.data.get("subject_filter", "timesheet")
        month_filter = request.data.get("month_filter", "")
        limit = min(to_int(request.data.get("limit"), default=10), 25)

        try:
            previews = fetch_mail_table_preview(mailbox, subject_filter, month_filter, limit)
        except Exception as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"previews": previews, "count": len(previews)})
